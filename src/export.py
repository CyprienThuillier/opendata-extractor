import csv
import json
import openpyxl as xl
import pandas as pd
import os

def export_file(targets: list, filename: str = "export.csv", append: bool = False):
    targets = [t for t in targets if t is not None]
    if not targets:
        return
    os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else ".", exist_ok=True)
    mode = "a" if append else "w"
    file_exists = append and os.path.exists(filename)
    with open(filename, mode, newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(targets[0].keys()))
        if not file_exists:
            writer.writeheader()
        writer.writerows(targets)

export_file_json = lambda targets, filename="export.json", append=False: export_file(targets, filename, append)

def export_file_excel(targets: list, filename: str = "export.xlsx", append: bool = False):
    targets = [t for t in targets if t is not None]
    if not targets:
        return
    os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else ".", exist_ok=True)
    if append and os.path.exists(filename):
        wb = xl.load_workbook(filename)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.append(list(targets[0].keys()))
        start_row = 2
    for target in targets:
        ws.append(list(target.values()))
    wb.save(filename)